from aiogram import Bot, Router, F
from aiogram.exceptions import TelegramAPIError
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from sqlalchemy.testing.config import any_async
from aiogram.enums import ParseMode
import asyncio
from Utils import get_calculated_period, filter_tickets_for_statistics
from datetime import datetime
from math import floor
from database.db import DataBase, redis_client
from colorama import Fore, Style
from aiogram.enums.parse_mode import ParseMode
from logger import logger
from core.dictionary import *
from handlers.Admin.keyboard.InlineKb import *
from config import *
from handlers.export import *
db = DataBase()
DATE_FORMAT = "%d.%m.%y"

#Временно, но скорее всего постоянно
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from aiogram.types import FSInputFile
from aiogram.filters import Filter
from sqlalchemy import select
from database.models import Users, Roles


admin_router = Router()
# 🔐 Фильтр доступа по роли admin (для всех сообщений)
class IsAdmin(Filter):
    async def __call__(self, message: Message) -> bool:
        async with db.Session() as session:
            result = await session.execute(
                select(Users, Roles.role_name)
                .join(Roles, Users.role_id == Roles.id)
                .where(Users.user_id == message.from_user.id)
            )
            row = result.first()
            return row is not None and row[1] == "admin"

# 🔐 Фильтр доступа по роли admin (для всех callback-кнопок)
class IsAdminCallback(Filter):
    async def __call__(self, call: CallbackQuery) -> bool:
        async with db.Session() as session:
            result = await session.execute(
                select(Roles.role_name)
                .join(Users, Users.role_id == Roles.id)
                .where(Users.user_id == call.from_user.id)
            )
            role_name = result.scalar_one_or_none()
            return role_name == "admin"

# ⛓️ Подключаем фильтры к admin_router
admin_router.message.filter(IsAdmin())
admin_router.callback_query.filter(IsAdminCallback())




class AddRoles(StatesGroup):
    add_name = State()


class AddService(StatesGroup):
    name_service = State()
    role_service = State()


class Banned(StatesGroup):
    add_id = State()
    delete_id = State()


class AddUserRole(StatesGroup):
    Username = State()
    role_user = State()

class StartMailing(StatesGroup):
    text_mailing = State()

class PaymentStates(StatesGroup):
    awaiting_new_rate = State()

CATEGORY_MAP = {
    "technical_support": "Техническая помощь / Technical Support",
    "payment_support": "Помощь с платежами / Payment Support",
    "hwid_reset": "NFA / HWID RESET",
    "reselling": "Reselling",
    "get_key": "Получить Ключ / Get a key",
    "bonus_per_50": "Бонус за каждые 50 тикетов"
}



@admin_router.message(Command(commands=['admin']), F.chat.type == "private")
async def admin(message: Message):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id} '
                            f'ввел команду /admin' + Style.RESET_ALL)
    try:
        if message.from_user.id == 434791099 or message.from_user.id == 835867765:
            await message.answer(text='Выберите действие: ', reply_markup=admin_panel)
        elif await db.get_admin_by_id(message.from_user.id) is True:
            await message.answer(text='Выберите действие:', reply_markup=admin_panel)
        else:
            await message.answer(text='Не известная команда, введите /start')
    except Exception as e:
        logger.error(f'Ошибка в команде /admin {e}')


@admin_router.callback_query(F.data.startswith('roles'))
async def all_roles(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал Роли' + Style.RESET_ALL)
    try:
        await call.message.edit_text(text='Выберите действие: ', reply_markup=admin_roles)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('back_menu'))
async def back_menu(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал Назад' + Style.RESET_ALL)
    try:
        await call.message.edit_text(text='Выберите действие: ', reply_markup=admin_panel)
    except Exception as e:
        logger.error(f'Ошибка в команде /back_menu {e}')


@admin_router.callback_query(F.data.startswith('black_list'))
async def black_list(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал Черный список' + Style.RESET_ALL)
    try:
        await call.message.edit_text(text='Выберите действие: ', reply_markup=admin_black_list)
    except Exception as e:
        logger.error(f'Ошибка в команде /black_list {e}')


@admin_router.callback_query(F.data.startswith('role_add'))
async def start_add_role(call: CallbackQuery, state: FSMContext):
    logger.info(
        Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id}  начал добавление роли' + Style.RESET_ALL)
    await call.message.delete()
    try:
        await call.message.answer(text='Введите название роли')
        await state.set_state(AddRoles.add_name)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.message(AddRoles.add_name)
async def add_role(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id} '
                            f'Ввел название роли: {message.text}' + Style.RESET_ALL)
    try:
        await state.update_data(role_name=message.text)
        result = await db.add_roles(message.text)
        if result is True:
            await message.answer(text='Роль добавлена', reply_markup=admin_panel)
            await state.clear()
        elif result == 'Name_is_occupied':
            await message.answer(text='Роль с таким названием уже существует, введите другое название',
                                 reply_markup=admin_panel)
            return
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('role_delete_'))
async def delete_role(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал удалить роль {call.data.split("_")[2]}' + Style.RESET_ALL)
    await call.message.delete()
    id_role = int(call.data.split("_")[2])
    try:
        result = await db.delete_roles(id_role)
        if result is True:
            await call.message.answer(text='Роль удалена', reply_markup=admin_panel)
        elif result == 'Роль не найдена!':
            await call.message.answer(text=result, reply_markup=admin_panel)
            return
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('role_del'))
async def delete_role(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал удалить роль' + Style.RESET_ALL)

    roles = await db.get_roles()
    roles_buttons = [InlineKeyboardButton(text=f'{role.role_name}', callback_data=f'role_delete_{role.id}') for role in
                     roles]
    back = [InlineKeyboardButton(text='🔙 Назад', callback_data='back_menu')]
    keyboard_buttons = InlineKeyboardMarkup(inline_keyboard=[[button] for button in roles_buttons] + [[back[0]]])
    try:
        await call.message.edit_text(text='Выберите роль, которую хотите удалить: ', reply_markup=keyboard_buttons)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('admin_service'))
async def services(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал Услуги' + Style.RESET_ALL)
    try:
        await call.message.edit_text(text='Выберите действие:', reply_markup=admin_services)
    except Exception as e:
        logger.error(f'Ошибка в команде /services {e}')


@admin_router.message(AddService.name_service)
async def add_service(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id} '
                            f'Ввел название услуги: {message.text}' + Style.RESET_ALL)
    roles = await db.get_roles()
    roles_buttons = [
        InlineKeyboardButton(text=f'{role.id}.{role.role_name}', callback_data=f'services_add_role_{role.id}') for
        role in roles]
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[button] for button in roles_buttons])
    try:
        await state.update_data(service_name=message.text)
        sent_message = await message.answer(text='Выберите роль или роли которые будут иметь доступ к услуге'
                                  'Пример ввода (1,2,3)', reply_markup=keyboard)


        await state.update_data(message_id=sent_message.message_id)
        await state.set_state(AddService.role_service)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.message(AddService.role_service)
async def add_service_role(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id}'
                            f'Выбрал роль для услуги: {message.text.strip()}' + Style.RESET_ALL)

    reg_data = await state.get_data()
    message_id = reg_data.get('message_id')
    if message_id:
        await bot.delete_message(
            chat_id=message.chat.id,
            message_id=message_id
        )
    service_name = reg_data.get('service_name')
    role_id = message.text.strip()
    try:
        result = await db.add_service(service_name, role_id)
        if result is True:
            await message.answer(text='Услуга добавлена', reply_markup=admin_panel)
            await state.clear()
        else:
            await message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('services_add'))
async def start_add_service(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал добавление услуги' + Style.RESET_ALL)

    await call.message.delete()
    try:
        await call.message.answer(text='Введите название услуги')
        await state.set_state(AddService.name_service)
    except Exception as e:
        logger.error(f'Ошибка {e}')


@admin_router.callback_query(F.data.startswith('services_del_'))
async def delete_service(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'выбрал удалить услугу {call.data.split("_")[2]}' + Style.RESET_ALL)
    await call.message.delete()
    service_id = int(call.data.split("_")[2])
    try:
        result = await db.service_delete(service_id)
        if result is True:
            await call.message.answer(text='Услуга удалена, выберите действие:', reply_markup=admin_panel)
        elif result == 'Услуга не найдена!':
            await call.message.answer(text=result, reply_markup=admin_panel)
            return
        else:
            logger.error(Fore.RED + f'{result}' + Style.RESET_ALL)
            await call.message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
    except Exception as e:
        logger.error(f'Ошибка {e}')
        await call.message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)


@admin_router.callback_query(F.data.startswith('services_del'))
async def start_del_service(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал удаление услуги' + Style.RESET_ALL)

    services = await db.get_services()
    services_buttons = [InlineKeyboardButton(text=f'{service.service_name}', callback_data=f'services_del_{service.id}')
                        for service in services]
    back = [InlineKeyboardButton(text='🔙 Назад', callback_data='back_menu')]
    keybord = InlineKeyboardMarkup(inline_keyboard=[[button] for button in services_buttons] + [[back[0]]])

    try:
        await call.message.edit_text(text='Выберите услугу, которую хотите удалить: ', reply_markup=keybord)
    except Exception as e:
        await call.message.answer("Ошибка")
        logger.error(Fore.RED + f'Ошибка {e}' + Style.RESET_ALL)


@admin_router.callback_query(F.data.startswith('blackList_add'))
async def start_add_black_list(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал добавление в черный список' + Style.RESET_ALL)
    await call.message.delete()
    try:
        await call.message.answer(text='Введите id пользователя, которого хотите добавить в черный список:')
        await state.set_state(Banned.add_id)
    except Exception as e:
        logger.error(Fore.RED + f'Ошибка {e}' + Style.RESET_ALL)


@admin_router.message(Banned.add_id)
async def add_black_list(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id}'
                            f'Ввел id пользователя: {message.text}' + Style.RESET_ALL)

    user_id = message.text.strip()

    try:
        result = await db.banned_users(user_id)
        if result is True:
            await message.answer(text='Пользователь добавлен в черный список', reply_markup=admin_panel)
            await state.clear()
        elif result == 'Пользователь уже в черном списке!':
            await message.answer(text=result, reply_markup=admin_panel)
            await state.clear()
        else:
            logger.error(Fore.RED + f'{result}' + Style.RESET_ALL)
            await message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
    except Exception as e:
        logger.error(Fore.RED + f'Ошибка {e}' + Style.RESET_ALL)
        await message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)


@admin_router.callback_query(F.data.startswith('blackList_del'))
async def start_del_black_list(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал удаление из черного списка' + Style.RESET_ALL)

    await call.message.delete()
    await call.message.answer(text='Введите id пользователя, которого хотите удалить из черного списка:')
    await state.set_state(Banned.delete_id)


@admin_router.message(Banned.delete_id)
async def del_black_list(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id} '
                            f'Ввел id пользователя: {message.text}' + Style.RESET_ALL)

    user_id = int(message.text.strip())
    try:
        result = await db.delete_banned_users(user_id)
        if result is True:
            await message.answer(text='Пользователь удален из черного списка', reply_markup=admin_panel)
            await state.clear()
        elif result == 'Пользователь не найден в черном списке!':
            await message.answer(text=result, reply_markup=admin_panel)
            await state.clear()
        else:
            await message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
            await state.clear()
    except Exception as e:
        logger.error(Fore.RED + f'Ошибка {e}' + Style.RESET_ALL)
        await message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
        await state.clear()


@admin_router.callback_query(F.data.startswith('roleUser_add'))
async def start_add_role_user(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал добавление роли пользователю' + Style.RESET_ALL)
    await call.message.delete()
    await call.message.answer(text='Введите @Username пользователя, которому хотите добавить роль:')
    await state.set_state(AddUserRole.Username)


@admin_router.message(AddUserRole.Username)
async def add_role_user(message: Message, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {message.from_user.username} id: {message.from_user.id} '
                            f'Ввел @Username пользователя: {message.text}' + Style.RESET_ALL)
    username = message.text.strip()

    if '@' in username:
        username = username.replace('@', '')

    users = await db.get_users_by_username(username)

    if users == 'Пользователь не найден!':
        await message.answer(text=users, reply_markup=admin_panel)
        await state.clear()
        return
    elif users is False:
        await message.answer(text='Ошибка, попробуйте позже!', reply_markup=admin_panel)
        await state.clear()
    else:
        await state.update_data(users=username)
        role = await db.get_roles()
        roles_buttons = [InlineKeyboardButton(text=f'{role.role_name}', callback_data=f'roleUser_user_{role.id}') for
                         role in role]
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[button] for button in roles_buttons])
        await message.answer(text='Выберите роль для пользователя:', reply_markup=keyboard)


@admin_router.callback_query(F.data.startswith('roleUser_user_'))
async def add_role_user(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'Выбрал роль для пользователя {call.data.split("_")[2]}' + Style.RESET_ALL)

    await call.message.delete()
    role_id = int(call.data.split("_")[2])
    reg_data = await state.get_data()
    username = reg_data.get('users')

    result = await db.add_user_role(username, role_id)
    if result is not False:
        await call.message.answer(text='Роль добавлена', reply_markup=admin_panel)
        await bot.send_message(
            chat_id=result['user_id'],
            text=f"✅ Вам выдана роль: {result['role_name']} 🎉\n"
                 "🔄 Чтобы у вас появились новые права, перезапустите бота: /start"
        )
        await state.clear()
    elif result == 'Пользователь не найден!':
        await call.message.answer(text=result, reply_markup=admin_panel)
        await state.clear()
    else:
        await call.message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
        await state.clear()


@admin_router.callback_query(F.data.startswith('role_user'))
async def start_del_role_user(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'отрыл меню выдачи роли' + Style.RESET_ALL)
    await call.message.edit_text(text='Выберите действие: ', reply_markup=admin_role_user_edit)


@admin_router.callback_query(F.data.startswith('roleUser_del_'))
async def start_del_role_user(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал удаление роли у пользователя {call.data.split("_")[2]}' + Style.RESET_ALL)

    user_id = int(call.data.split("_")[2])
    await call.message.delete()
    result = await db.delete_user_role(user_id)
    if result is True:
        await call.message.answer(text='Роль удалена', reply_markup=admin_panel)
        message_text = ('🔔 Уведомление 🔔\n\n'
                   '📢 Вы сняты с должности.'
                   '🙏 Благодарим вас за проделанную работу! Ваш вклад не остался незамеченным.'
                   '✨ Желаем вам успехов в будущем, новых достижений и всего самого доброго! 🌟'
                   'С уважением, команда 🦊GAMEBREAKER 🤝'
        )
        await bot.send_message(chat_id=user_id, text=message_text)
        await state.clear()
    else:
        await call.message.answer(text='Ошибка, попробуйте позже', reply_markup=admin_panel)
        await state.clear()


@admin_router.callback_query(F.data.startswith('roleUser_del'))
async def start_del_role_user(call: CallbackQuery, state: FSMContext):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал удаление роли у пользователя' + Style.RESET_ALL)

    users = await db.get_user_role()
    if users is False:
        await call.message.answer(text='Ошибка, попробуйте позже!', reply_markup=admin_panel)
        await state.clear()
        return
    elif users == 'Пользователь не найден!':
        await call.message.answer(text=users, reply_markup=admin_panel)
        await state.clear()
    else:
        user_buttons = [InlineKeyboardButton(text=f'{user.username}', callback_data=f'roleUser_del_{user.user_id}')
                        for user in users]
        back = [InlineKeyboardButton(text='🔙 Назад', callback_data='role_user')]
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[button] for button in user_buttons] + [[back[0]]])
        await call.message.edit_text(text='Выберите пользователя, у которого хотите удалить роль:',
                                     reply_markup=keyboard)


@admin_router.callback_query(F.data.startswith('export'))
async def start_export(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал экспорт' + Style.RESET_ALL)
    await call.message.delete()
    await export_data(call, bot)

# --- Кнопки подтверждения ---
confirm_mailing_kb = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Начать рассылку", callback_data="confirm_real_send")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_mailing")]
    ]
)
# Админ нажал "Начать рассылку"
@admin_router.callback_query(F.data.startswith('malling_message'))
async def start_mailing(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} хочет начать рассылку' + Style.RESET_ALL)

    # 🧹 Очистим FSM
    await state.clear()

    # 🧼 Попробуем удалить последние 5 сообщений (где могли быть кнопки)
    for msg_id in range(call.message.message_id - 1, call.message.message_id - 6, -1):
        try:
            await bot.edit_message_reply_markup(
                chat_id=call.message.chat.id,
                message_id=msg_id,
                reply_markup=None
            )
        except Exception:
            pass  # Игнорируем ошибки если сообщение уже не редактируется

    # 🧼 Удалим текущие кнопки
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception as e:
        logger.warning(f"Не удалось удалить клавиатуру текущего сообщения: {e}")

    # ✉️ Запрос сообщения
    await call.message.answer("✉️ Пришли сообщение, которое хочешь разослать:")
    await state.set_state(StartMailing.text_mailing)



# Получено сообщение, ждём подтверждения
@admin_router.message(StartMailing.text_mailing)
async def receive_mailing_text(message: Message, state: FSMContext, bot):
    await state.update_data(message_id=message.message_id, chat_id=message.chat.id)
    await message.answer("\u2705 Сообщение получено. Начинаем рассылку?", reply_markup=confirm_mailing_kb)

    user = message.from_user
    await bot.send_message(
        434791099,
        f"⚠️ Пользователь @{user.username or 'без username'} (id: {user.id}) отправил сообщение для рассылки и ожидает подтверждения."
    )

# Подтверждена рассылка
@admin_router.callback_query(F.data == "confirm_real_send")
async def do_real_mailing(call: CallbackQuery, state: FSMContext, bot):
    # ⬇ Удаляем кнопки подтверждения сразу
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception as e:
        logger.warning(f"Не удалось удалить кнопки подтверждения: {e}")
    data = await state.get_data()
    message_id = data.get("message_id")
    from_chat_id = data.get("chat_id")

    # 🛡️ Проверка на наличие данных перед запуском
    if not message_id or not from_chat_id:
        await call.message.edit_text(
            "❌ Ошибка: не удалось найти сообщение для рассылки. "
            "Возможно, вы вызвали меню повторно или состояние было сброшено.\n"
            "Попробуйте начать заново."
        )
        await state.clear()
        return

    await state.clear()

    users = await db.get_user_all()
    user_ids = [user.user_id for user in users]

    if not user_ids:
        await call.message.edit_text("❌ Нет пользователей для рассылки!")
        return

    await bot.send_message(
        434791099,
        f"⚠️ Пользователь @{call.from_user.username or 'без username'} (id: {call.from_user.id}) подтвердил рассылку. Начинаем отправку..."
    )

    success = 0
    errors = 0
    total = len(user_ids)

    progress = await call.message.answer(f"⌛  Прогресс: 0/{total}")

    for index, user_id in enumerate(user_ids, 1):
        try:
            await bot.copy_message(
                chat_id=user_id,
                from_chat_id=from_chat_id,
                message_id=message_id
            )
            success += 1
        except Exception as e:
            if "Too Many Requests" in str(e):
                await progress.edit_text("(⚠ Лимит! Ждём 10 сек...")
                await asyncio.sleep(10)
                continue
            errors += 1
            logger.error(f"Ошибка {user_id}: {e}")

        if index % 10 == 0:
            await progress.edit_text(f"⌛ Прогресс: {index}/{total}")

        await asyncio.sleep(0.1)

    await progress.delete()
    await call.message.answer(
        f"✅ Рассылка завершена!\n\n"
        f"✅ Успешно: {success}\n"
        f"❌ Ошибок: {errors}\n\n"
        f"⚠️ Всего: {total}"
    )

    await bot.send_message(
        434791099,
        f"✅ Рассылка завершена.\n\n"
        f"✉ Отправитель: @{call.from_user.username or 'без username'} ({call.from_user.id})"
    )

# Отменена рассылка
@admin_router.callback_query(F.data == "cancel_mailing")
async def cancel_mailing(call: CallbackQuery, state: FSMContext):
    # ⬇ Удаляем кнопки
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception as e:
        logger.warning(f"Не удалось удалить кнопки отмены: {e}")

    await state.clear()
    await call.message.edit_text("❌ Рассылка отменена.")




@admin_router.callback_query(F.data.startswith('start_media'))
async def start_media(call: CallbackQuery, state: FSMContext, bot: Bot):
    logger.info(Fore.BLUE + f'Пользователь {call.from_user.username} id: {call.from_user.id} '
                            f'начал загрузку медиа' + Style.RESET_ALL)

    await call.message.edit_text(text='Выберите действие:', reply_markup=admin_media_menu_kb)


@admin_router.callback_query(F.data == "edit_rates")
async def show_users_for_rate_edit(callback: CallbackQuery):
    users = await db.get_users_with_roles_for_rates()
    if not users:
        await callback.message.edit_text("Нет пользователей с правом настройки ставок.")
        return

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
                            [InlineKeyboardButton(text=f"@{username}", callback_data=f"edit_rate:{user_id}")]
                            for user_id, username in users
                        ] + [[InlineKeyboardButton(text="⬅️ Назад", callback_data="back_menu")]]
    )
    await callback.message.edit_text("Выберите пользователя для настройки ставок:", reply_markup=kb)

@admin_router.callback_query(F.data.startswith("edit_rate:"))
async def show_categories(callback: CallbackQuery):
    user_id = int(callback.data.split(":")[1])
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
                            [InlineKeyboardButton(text=cat, callback_data=f"set_rate:{user_id}:{field}")]
                            for field, cat in CATEGORY_MAP.items()
                        ] + [[InlineKeyboardButton(text="⬅️ Назад", callback_data="back_to_user_list")]]
    )
    await callback.message.edit_text("Выберите категорию для изменения ставки:", reply_markup=kb)

@admin_router.callback_query(F.data.startswith("set_rate:"))
async def prompt_new_rate(callback: CallbackQuery, state: FSMContext):
    _, user_id, field = callback.data.split(":")
    user = await db.get_user_by_id(int(user_id))
    record = await db.get_payment_rate(int(user_id))
    if record is None:
        await db.create_payment_rate(int(user_id))
        record = await db.get_payment_rate(int(user_id))

    current = getattr(record, field)

    await state.update_data(user_id=int(user_id), field=field)
    await callback.message.edit_text(
        f"Текущая ставка: {current} руб. для @{user.username} по категории:\n"
        f"\"{CATEGORY_MAP[field]}\"\n\nВведите новое значение:"
    )
    await state.set_state(PaymentStates.awaiting_new_rate)

@admin_router.callback_query(F.data == "back_to_user_list")
async def back_to_user_list(callback: CallbackQuery):
    users = await db.get_users_with_roles_for_rates()
    if not users:
        await callback.message.edit_text("Нет доступных пользователей.")
        return

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
                            [InlineKeyboardButton(text=f"@{username}", callback_data=f"edit_rate:{user_id}")]
                            for user_id, username in users
                        ] + [[InlineKeyboardButton(text="⬅️ Назад", callback_data="back_menu")]]
    )
    await callback.message.edit_text("Выберите пользователя:", reply_markup=keyboard)

@admin_router.message(PaymentStates.awaiting_new_rate)
async def save_new_rate(message: Message, state: FSMContext):
    data = await state.get_data()
    user_id = data["user_id"]
    field = data["field"]
    if not field:
        await message.answer("⚠️ Ошибка: не выбрана категория.")
        await state.clear()
        return

    # Преобразуем ввод в число
    try:
        value = int(message.text)
    except ValueError:
        await message.answer("❌ Введите корректное целое число.")
        return

    # Ограничения: не менее 0 и не более 15 000
    if value < 0 or value > 15000:
        await message.answer("⚠️ Значение должно быть от 0 до 15 000 руб.")
        return

    # Получаем username
    username = await db.get_username_by_id(user_id)
    if not username:
        await message.answer("⚠️ Пользователь не найден.")
        await state.clear()
        return

    if field == "technical_support" and user_id == 434791099 and value < 80:
        await message.answer("⚠️ Ставка для @jarkadash по Technical Support не может быть ниже 80 руб.")
        return

    # Получаем/создаём запись
    record = await db.get_payment_rate(user_id)
    if not record:
        await db.create_payment_rate(user_id)

    # Обновляем ставку
    await db.update_payment_rate(user_id, field, value)

    # Название категории
    category = CATEGORY_MAP.get(field, field)

    await message.answer(
        f"✅ Ставка @{username} для категории \"{category}\" успешно обновлена: {value} руб.",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ Назад к категориям", callback_data=f"edit_rate:{user_id}")]
            ]
        )
    )

    await state.clear()


@admin_router.message(F.text.startswith("/allstats"))
async def all_stats_command(message: Message, bot: Bot):
    try:
        import re
        from datetime import datetime

        raw = message.text.strip()

        if raw.startswith("/allstats"):
            raw = raw[len("/allstats"):].strip()

        if not raw:
            await message.answer("📅 Укажите диапазон в формате:\n`/allstats 11.07.25–25.07.25`", parse_mode="Markdown")
            start_date, end_date = get_calculated_period()
        else:
            parts = re.split(r"\s*[–—\-~]\s*", raw)
            if len(parts) != 2:
                await message.answer("❗ Формат: `/allstats ДД.ММ.ГГ–ДД.ММ.ГГ`", parse_mode="Markdown")
                return

            left, right = parts[0].strip(), parts[1].strip()

            date_formats = []
            try:
                date_formats.append(DATE_FORMAT)
            except NameError:
                pass
            date_formats.extend(["%d.%m.%Y", "%d.%m.%y"])

            def try_parse(s: str):
                for fmt in date_formats:
                    try:
                        return datetime.strptime(s, fmt).date()
                    except ValueError:
                        continue
                return None

            right_date = try_parse(right)
            if right_date is None:
                await message.answer(
                    "❗ Не удалось распознать правую дату. Пример: `/allstats 11.07.25–25.07.25`",
                    parse_mode="Markdown"
                )
                return

            if re.fullmatch(r"\d{2}\.\d{2}", left):
                left = f"{left}.{right_date.year}"

            left_date = try_parse(left)
            if left_date is None:
                await message.answer(
                    "❗ Не удалось распознать левую дату. Примеры:\n"
                    "`/allstats 11.07–25.07.25`\n"
                    "`/allstats 11.07.25–25.07.25`",
                    parse_mode="Markdown"
                )
                return

            if right_date < left_date:
                left_date, right_date = right_date, left_date

            start_date, end_date = left_date, right_date

        user = message.from_user
        await bot.send_message(
            434791099,
            f"ℹ️ Пользователь @{user.username or 'без username'} (id: {user.id}) воспользовался командой /allstats"
        )

        users = await db.get_users_with_roles_for_rates()
        if not users:
            await message.answer("Нет сотрудников с ролью support/admin.")
            return

        from Utils import filter_tickets_for_statistics

        text = f"📊 Статистика с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}\n\n"

        all_ticket_rows = []

        async with db.Session() as session:
            for row in users:
                user_id = row[0]
                username = row[1] or "без username"

                included, excluded = await filter_tickets_for_statistics(session, user_id, start_date, end_date)

                def ticket_to_row(ticket, excluded_reason=None):
                    return {
                        "id": ticket.id,
                        "client_id": ticket.client_id,
                        "client_name": ticket.client_name,
                        "support_id": ticket.support_id,
                        "support_name": ticket.support_name,
                        "service_id": ticket.service_id,
                        "service_name": ticket.service_name,
                        "created_at": ticket.created_at,
                        "accept_at": ticket.accept_at,
                        "completed_at": ticket.completed_at,
                        "status": ticket.status,
                        "stars": ticket.stars,
                        "description": ticket.description,
                        "excluded_reason": excluded_reason
                    }

                for ticket in included:
                    all_ticket_rows.append(ticket_to_row(ticket))
                for ticket, reason in excluded:
                    all_ticket_rows.append(ticket_to_row(ticket, excluded_reason=reason))

        df = pd.DataFrame(all_ticket_rows)

        CATEGORY_ORDER = {
            "Техническая помощь / Technical Support": "Техническая помощь",
            "Помощь с платежами / Payment Support": "Помощь с платежами",
            "NFA / HWID RESET": "NFA/HWID",
            "Reselling": "Reselling",
            "Получить Ключ / Get a key": "Выдача ключей"
        }

        async with db.Session() as session:
            for row in users:
                user_id = row[0]
                username = row[1] or "без username"

                user_df = df[
                    (df["support_id"] == user_id) &
                    (
                            (df["excluded_reason"].isnull()) |
                            (df["excluded_reason"].astype(str).str.strip() == "")
                    )
                    ]

                if user_df.empty:
                    continue

                counts = user_df["service_name"].value_counts().to_dict()

                rates = await db.get_user_rates(session, user_id)
                salary = 0
                total = sum(counts.values())

                for service, count in counts.items():
                    rate = rates.get(service, 0)

                    if service == "Техническая помощь / Technical Support" and user_id == 434791099 and rate < 80:
                        rate = 80

                    salary += count * rate

                bonus = rates.get("Бонус", 0)
                if bonus and total >= 50:
                    salary += floor(total / 50) * bonus

                text += f"👨‍💻 @{username}:\n"
                for category in CATEGORY_ORDER:
                    if category in counts:
                        text += f"- {CATEGORY_ORDER[category]}: {counts[category]}\n"
                text += f"🧾 Всего: {total}\n"
                formatted_salary = f"{salary:,.0f}".replace(",", " ")
                text += f"💰 Заработал: {formatted_salary} руб.\n\n"

        if message.from_user.id == 434791099 and message.chat.type == "private" and not df.empty:
            filename = f"Отчет_allstats_{start_date.strftime('%d.%m.%y')}_{end_date.strftime('%d.%m.%y')}.xlsx"
            df.to_excel(filename, index=False)

            time.sleep(0.2)

            try:
                wb = load_workbook(filename)
            except FileNotFoundError:
                await message.answer("❗ Не удалось открыть Excel-файл: он не был создан.")
                return
            except PermissionError:
                await message.answer("❗ Файл уже открыт в Excel. Пожалуйста, закройте его и повторите.")
                return

            ws = wb.active
            fill = PatternFill(start_color="C7C385", end_color="C7C385", fill_type="solid")
            excluded_col_index = list(df.columns).index("excluded_reason") + 1

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[excluded_col_index - 1].value:
                    for cell in row:
                        cell.fill = fill

            wb.save(filename)
            logger.info("[ALLSTATS] Excel-файл успешно создан: %s", filename)

            try:
                file = FSInputFile(filename)
                await message.answer_document(file, caption="📎 Excel отчет")
                os.remove(filename)
                logger.info("[ALLSTATS] Excel-файл отправлен и удалён")
            except Exception as e:
                logger.error(f"[ALLSTATS] Ошибка при отправке/удалении файла: {e}", exc_info=True)
                await message.answer("⚠️ Не удалось отправить Excel-файл.")
        else:
            reason_parts = []
            if message.from_user.id != 434791099:
                reason_parts.append(f"user_id ≠ A ({message.from_user.id})")
            if message.chat.type != "private":
                reason_parts.append(f"chat_type ≠ private ({message.chat.type})")
            if df.empty:
                reason_parts.append("нет данных (df пуст)")
            logger.info("[ALLSTATS] Excel-файл не создан: " + "; ".join(reason_parts))

        await message.answer(text.strip())

    except Exception as e:
        logger.error(f"[ALLSTATS ERROR] {e}", exc_info=True)
        await message.answer("❗ Ошибка при обработке команды /allstats.")

JKEYID = 434791099
AUTOACCEPT_ENABLED_KEY = "autoaccept:enabled"
AUTOACCEPT_OWNER_KEY = "autoaccept:owner"

@admin_router.message(Command("autoaccept_on"), F.chat.type == "private")
async def _autoaccept_on(message: Message):
    if message.from_user.id != JKEYID:
        return
    await redis_client.set(AUTOACCEPT_ENABLED_KEY, "1")
    await redis_client.set(AUTOACCEPT_OWNER_KEY, str(JKEYID))
    await message.answer("АП ✅")

@admin_router.message(Command("autoaccept_off"), F.chat.type == "private")
async def _autoaccept_off(message: Message):
    if message.from_user.id != JKEYID:
        return
    await redis_client.set(AUTOACCEPT_ENABLED_KEY, "0")
    await redis_client.delete(AUTOACCEPT_OWNER_KEY)
    await message.answer("АП ⛔")

@admin_router.message(Command("autoaccept_status"), F.chat.type == "private")
async def _autoaccept_status(message: Message):
    if message.from_user.id != JKEYID:
        return
    enabled = await redis_client.get(AUTOACCEPT_ENABLED_KEY)
    owner = await redis_client.get(AUTOACCEPT_OWNER_KEY)
    active_cnt = await db.count_active_for(JKEYID)
    await message.answer(f"enabled: {enabled or '0'}\nowner: {owner or '-'}\nactive tickets: {active_cnt}")

